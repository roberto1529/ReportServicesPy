from openpyxl import load_workbook
import main as arranque
# Cargar la plantilla
def genearReporte():
    workbook = load_workbook('./templates/excel/datos.xlsx', read_only=False, data_only=True)  # Asegurarse de leer los valores, no las fórmulas
    sheet = workbook.active
    versionDocto = '1.0.0';
    buisness = 'ONECORE SAS'
    
    # Datos de ejemplo
    datos = [
        {'nombre': 'Juan', 'edad': 30},
        {'nombre': 'María', 'edad': 25}
    ]

    # Fila inicial para escribir los datos (A10)
    fila_inicial = 9

    # Recorrer las filas y reemplazar los marcadores
    for i, dato in enumerate(datos):
        fila = fila_inicial + i
        for col, key in enumerate(['nombre', 'edad'], start=1):
            cell = sheet.cell(row=fila, column=col)
            cell.value = dato.get(key, '')  # Usar get para manejar claves inexistentes
            # Opcional: Formatear la celda
            # cell.number_format = 'General'  # Ajusta según tu formato deseado
    
    # Buscar y reemplazar el marcador de posición
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == '{{version}}':
                cell.value = versionDocto
            if cell.value == '{{empresa}}':
                cell.value = buisness

    # Guardar el archivo
    workbook.save('./outputs/workbook.xlsx')

def rootSystem():
    arranque.limpiarReportes()
    genearReporte()

if __name__ == '__main__':
    rootSystem()